"""XLSX-specific chunker for RAG ingestion.

Converts each sheet in an Excel workbook to CSV, then delegates to the
existing row-based CSV chunker. This avoids Docling's slow and
memory-intensive table parsing while preserving header-per-chunk structure
that produces better embeddings for tabular data.
"""

import csv
import io
import logging
from typing import List

from openpyxl import load_workbook

from .csv_chunker import chunk_csv

logger = logging.getLogger(__name__)


def _is_likely_header(row: tuple, next_row: tuple = None) -> bool:
    """
    Determine if a row looks like a column header.

    Uses two signals:
    1. The row is predominantly non-numeric text strings
    2. If a next row is available, it should look different (contain numbers/dates),
       confirming this row is labels and the next is data

    Returns False for rows that are mostly empty (title/banner rows).
    """
    non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
    total_cells = len(row) if row else 0

    # Skip rows where most cells are empty (title/banner rows)
    if total_cells > 0 and len(non_empty) < total_cells * 0.5:
        return False

    if not non_empty:
        return False

    # Check if most cells are non-numeric text
    text_count = 0
    for cell in non_empty:
        if isinstance(cell, str):
            stripped = cell.strip()
            try:
                float(stripped.replace(",", ""))
                continue  # It's a number in string form
            except ValueError:
                pass
            text_count += 1
        # Non-string types (int, float, datetime) are not header-like

    if text_count <= len(non_empty) * 0.7:
        return False

    # If we have a next row, verify it looks like data (has numbers or dates)
    if next_row is not None:
        next_non_empty = [cell for cell in next_row if cell is not None and str(cell).strip()]
        numeric_count = 0
        for cell in next_non_empty:
            if isinstance(cell, (int, float)):
                numeric_count += 1
            elif isinstance(cell, str):
                try:
                    float(cell.strip().replace(",", ""))
                    numeric_count += 1
                except ValueError:
                    pass
        # Data row should have at least some numbers
        if next_non_empty and numeric_count > 0:
            return True
        # If next row is also all text, this might not be the header
        if next_non_empty and numeric_count == 0:
            return False

    return True


def _find_header_row_index_from_rows(rows: list) -> int:
    """
    Find the first row that looks like a real column header.

    A header row is one where:
    - Most cells are populated (not a sparse title row)
    - Most cells are text strings (not numbers/dates)
    - The following row contains data (numbers, dates, mixed types)

    Falls back to row 0 if no clear header is found.

    Args:
        rows: List of non-empty row tuples.

    Returns:
        0-based index of the header row within the list.
    """
    if len(rows) <= 1:
        return 0

    for i, row in enumerate(rows[:10]):
        next_row = rows[i + 1] if i + 1 < len(rows) else None
        if _is_likely_header(row, next_row):
            if i > 0:
                logger.info(f"Skipping {i} non-header row(s) before detected header")
            return i

    return 0


def chunk_xlsx(file_bytes: bytes, max_tokens: int = 900) -> List[str]:
    """
    Chunk an XLSX file by converting each sheet to CSV and chunking rows.

    Each sheet is processed independently. The sheet name is prepended as
    context to every chunk from that sheet so the embedding captures which
    sheet the data belongs to.

    Args:
        file_bytes: Raw XLSX file content.
        max_tokens: Maximum token count per chunk (passed to chunk_csv).

    Returns:
        List of text chunks across all sheets.
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    all_chunks: List[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Collect non-empty rows (read_only mode only allows one iteration)
        non_empty_rows = []
        for row in ws.iter_rows(values_only=True):
            if not all(cell is None for cell in row):
                non_empty_rows.append(row)

        if not non_empty_rows:
            logger.info(f"Sheet '{sheet_name}' is empty, skipping")
            continue

        # Detect real header row (skip title/banner rows)
        header_offset = _find_header_row_index_from_rows(non_empty_rows)
        data_rows = non_empty_rows[header_offset:]

        # Convert to CSV
        buf = io.StringIO()
        writer = csv.writer(buf)
        for row in data_rows:
            writer.writerow([str(cell) if cell is not None else "" for cell in row])

        csv_bytes = buf.getvalue().encode("utf-8")
        sheet_chunks = chunk_csv(csv_bytes, max_tokens=max_tokens)

        # Prepend sheet name for multi-sheet context
        if len(wb.sheetnames) > 1:
            sheet_chunks = [f"Sheet: {sheet_name}\n{chunk}" for chunk in sheet_chunks]

        logger.info(f"Sheet '{sheet_name}': {len(data_rows)} rows -> {len(sheet_chunks)} chunks")
        all_chunks.extend(sheet_chunks)

    wb.close()
    logger.info(f"XLSX chunked into {len(all_chunks)} total chunks across {len(wb.sheetnames)} sheets")
    return all_chunks
