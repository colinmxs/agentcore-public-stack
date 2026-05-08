"""Analyze spreadsheet files using Code Interpreter.

Factory function creates a context-bound tool that downloads tabular files
from S3, pushes them to Code Interpreter, and executes Python code for analysis.
"""

import logging
import os
import re
from typing import Any, Dict, Optional

import boto3
from strands import tool

from .list_spreadsheets_tool import _get_kb_files, _get_session_files

logger = logging.getLogger(__name__)

MAX_OUTPUT_CHARS = 10000  # ~2500 tokens — safe margin under context limits
MAX_ERROR_CHARS = 600  # cleaned traceback budget — full pandas tracebacks are noise

_SCHEMA_MARKER = "[__SCHEMA__]"
_SHEETS_MARKER = "[__SHEETS__]"


def _truncate_output(text: str) -> str:
    """Truncate tool output to prevent blowing the LLM context window."""
    if not text or len(text) <= MAX_OUTPUT_CHARS:
        return text
    return text[:MAX_OUTPUT_CHARS] + f"\n\n... (output truncated — {len(text):,} chars total, showing first {MAX_OUTPUT_CHARS:,})"


def _strip_first_row(schema: str) -> str:
    """Drop the ``first_row: ...`` line from a schema footer.

    On the happy path the first-row preview helps the model write correct
    code. On the error path the model already has the load line and column
    list — the full row dump is ~30 fields of noise. This trims it.
    """
    return "\n".join(
        line for line in schema.splitlines()
        if not line.startswith("first_row:")
    )


# ---------------------------------------------------------------------------
# Stderr cleaning
# ---------------------------------------------------------------------------

# Frames we never want to show the LLM — they're pandas/numpy internals with
# zero signal for fixing the user's code.
_INTERNAL_FRAME_MARKERS = (
    "site-packages/pandas/",
    "site-packages/numpy/",
    "pandas/_libs/",
    "pandas/core/",
    "pandas/io/",
)


def _clean_stderr(stderr: str) -> str:
    """Strip pandas internal frames and dtype warnings from a traceback.

    Keeps the user-code frame (the `/tmp/ipykernel_*.py` line they wrote) and
    the final exception line. Falls back to a truncated raw stderr if the
    traceback doesn't match the expected shape.
    """
    if not stderr:
        return "Unknown error"

    lines = stderr.splitlines()

    # 1. Drop DtypeWarning noise (spans 2 lines: the warning + the call-site).
    filtered: list[str] = []
    skip_next = False
    for line in lines:
        if skip_next:
            skip_next = False
            continue
        if "DtypeWarning:" in line or "FutureWarning:" in line or "UserWarning:" in line:
            skip_next = True  # next line is usually the offending code snippet
            continue
        filtered.append(line)

    # 2. Find the final exception line (e.g. "KeyError: 'NET_AMOUNT'").
    final_exception = ""
    for line in reversed(filtered):
        stripped = line.strip()
        if not stripped:
            continue
        # Exception lines are left-flush and match "ExceptionName: message".
        if not line.startswith((" ", "\t")) and re.match(r"^[A-Z][A-Za-z]*(?:Error|Exception|Warning):", stripped):
            final_exception = stripped
            break

    # 3. Find the user-code frame (ipykernel tempfile, not site-packages).
    user_frame_lines: list[str] = []
    for i, line in enumerate(filtered):
        stripped = line.strip()
        if not stripped.startswith("File "):
            continue
        if any(m in stripped for m in _INTERNAL_FRAME_MARKERS):
            continue
        # Keep this frame + up to the next 2 lines (the code snippet + pointer).
        user_frame_lines.append(stripped)
        for j in range(i + 1, min(i + 3, len(filtered))):
            nxt = filtered[j].strip()
            if not nxt or nxt.startswith("File "):
                break
            user_frame_lines.append(nxt)
        break

    if user_frame_lines and final_exception:
        cleaned = "\n".join(user_frame_lines) + "\n" + final_exception
    elif final_exception:
        cleaned = final_exception
    else:
        # Unrecognized shape — return a short tail rather than a 3K dump.
        cleaned = "\n".join(filtered[-8:]).strip()

    if len(cleaned) > MAX_ERROR_CHARS:
        cleaned = cleaned[:MAX_ERROR_CHARS] + " ..."
    return cleaned


# ---------------------------------------------------------------------------
# Schema preview probe
# ---------------------------------------------------------------------------


def _build_preview_code(csv_filename: str) -> str:
    """Return Python code that prints a compact schema snapshot for csv_filename.

    Runs a bounded skiprows probe (0..8) to handle report-style exports with
    leading metadata rows. Picks the skiprows value that produces the cleanest
    header — no ``Unnamed:`` columns, no duplicates, non-empty names — and
    emits a ready-to-use ``pd.read_csv(...)`` invocation when the best
    candidate is meaningfully better than skiprows=0. Otherwise it reports the
    columns at skiprows=0 and lets the model decide.

    Output is bracketed with _SCHEMA_MARKER so it can be reliably extracted
    from the interpreter's stdout stream even if user code prints other things.
    """
    return f"""
import warnings, pandas as pd
warnings.filterwarnings('ignore')

def _score(cols):
    # Higher is better. Punishes Unnamed columns and duplicates.
    if not cols:
        return -10_000
    unnamed = sum(1 for c in cols if str(c).startswith('Unnamed:'))
    named = len(cols) - unnamed
    dup_penalty = (len(cols) - len(set(cols))) * 20
    blank_penalty = sum(1 for c in cols if not str(c).strip()) * 10
    return named - (unnamed * 5) - dup_penalty - blank_penalty

try:
    with open({csv_filename!r}, 'r') as _fh:
        _total_rows = sum(1 for _ in _fh)

    # Score skiprows=0..8, keep the winner and remember the baseline.
    _baseline_score, _baseline_cols = -float('inf'), []
    _best_skip, _best_score, _best_cols = 0, -float('inf'), []
    for _sk in range(9):
        try:
            _cols = pd.read_csv({csv_filename!r}, nrows=0, skiprows=_sk, low_memory=False).columns.tolist()
        except Exception:
            continue
        _sc = _score(_cols)
        if _sk == 0:
            _baseline_score, _baseline_cols = _sc, _cols
        if _sc > _best_score:
            _best_skip, _best_score, _best_cols = _sk, _sc, _cols

    # Confidence gate: only prescribe a non-zero skiprows when the winner
    # actually fixes a header problem — either more named columns OR fewer
    # Unnamed columns than the baseline — AND the winner is mostly clean.
    # A score-delta threshold alone can't distinguish "found the real header"
    # from "data row happens to parse cleanly", so we anchor on named/unnamed
    # counts instead.
    def _named_unnamed(cols):
        u = sum(1 for c in cols if str(c).startswith('Unnamed:'))
        return len(cols) - u, u
    _base_named, _base_unnamed = _named_unnamed(_baseline_cols)
    _win_named, _win_unnamed = _named_unnamed(_best_cols)
    _win_clean_ratio = _win_named / max(len(_best_cols), 1)

    _prescribe = (
        _best_skip > 0
        and _win_clean_ratio >= 0.7
        and (_win_named > _base_named or _win_unnamed < _base_unnamed)
    )

    if _prescribe:
        _report_skip, _report_cols = _best_skip, _best_cols
    else:
        _report_skip, _report_cols = 0, _baseline_cols

    _data_rows = max(_total_rows - 1 - _report_skip, 0)
    _col_preview = ', '.join(str(c) for c in _report_cols[:20])
    if len(_report_cols) > 20:
        _col_preview += f' ... (+{{len(_report_cols) - 20}} more)'

    try:
        _head = pd.read_csv({csv_filename!r}, skiprows=_report_skip, nrows=1, low_memory=False)
        _first_row = _head.iloc[0].to_dict() if len(_head) else {{}}
        _first_row = {{k: (str(v)[:40] + '...' if len(str(v)) > 40 else v) for k, v in _first_row.items()}}
    except Exception:
        _first_row = {{}}

    if _prescribe:
        _load = f"pd.read_csv({csv_filename!r}, skiprows={{_report_skip}}, low_memory=False)"
        _note = f"  # {{_report_skip}} metadata row(s) detected before header"
    else:
        _load = f"pd.read_csv({csv_filename!r}, low_memory=False)"
        _note = ""

    print({_SCHEMA_MARKER!r})
    print(f'file: {csv_filename} ({{_data_rows}} rows x {{len(_report_cols)}} cols)')
    print(f'load: {{_load}}{{_note}}')
    print(f'columns: {{_col_preview}}')
    print(f'first_row: {{_first_row}}')
    # If confidence was low, flag it so the model knows to verify.
    if not _prescribe and _win_unnamed > 0 and _win_unnamed < len(_best_cols):
        print(f'note: header may need adjustment (skiprows=0 has {{_base_unnamed}}/{{len(_baseline_cols)}} unnamed columns); inspect head() if unsure')
    print({_SCHEMA_MARKER!r})
except Exception as _e:
    print({_SCHEMA_MARKER!r})
    print(f'schema preview unavailable: {{_e}}')
    print({_SCHEMA_MARKER!r})
"""


def _extract_schema_preview(stdout: str) -> tuple[str, str]:
    """Split stdout into (schema_block, remaining_stdout).

    The schema block is whatever is between _SCHEMA_MARKER pairs; if no markers
    are found, returns ("", stdout).
    """
    if _SCHEMA_MARKER not in stdout:
        return "", stdout
    parts = stdout.split(_SCHEMA_MARKER)
    # parts = [before, schema, after, ...]; stitch back everything non-schema.
    if len(parts) >= 3:
        schema = parts[1].strip()
        remaining = (parts[0] + _SCHEMA_MARKER.join(parts[2:])).strip("\n")
        return schema, remaining
    return "", stdout


def _get_code_interpreter_id() -> Optional[str]:
    """Get Code Interpreter ID from environment or SSM."""
    ci_id = os.getenv("AGENTCORE_CODE_INTERPRETER_ID")
    if ci_id:
        return ci_id
    try:
        project_name = os.getenv("PROJECT_NAME", "strands-agent-chatbot")
        environment = os.getenv("ENVIRONMENT", "dev")
        region = os.getenv("AWS_REGION", "us-west-2")
        ssm = boto3.client("ssm", region_name=region)
        response = ssm.get_parameter(Name=f"/{project_name}/{environment}/agentcore/code-interpreter-id")
        return response["Parameter"]["Value"]
    except Exception:
        return None


def make_analyze_tool(
    assistant_id: Optional[str],
    session_id: str,
    user_id: str,
):
    """Create an analyze_spreadsheet tool bound to the given context."""

    @tool
    def analyze_spreadsheet(
        filename: str,
        python_code: str,
        output_filename: Optional[str] = None,
    ) -> Dict[str, Any]:
        """Analyze a spreadsheet file using Python code in Code Interpreter.

        Downloads the specified file and loads it into a sandboxed Python
        environment for analysis. Use pandas, numpy, matplotlib, and seaborn.

        ⚠️  CRITICAL — filename vs. in-sandbox path
        -------------------------------------------
        The ``filename`` parameter names the **source** file (exactly as it
        appears in the chat attachment or knowledge base, e.g.
        ``"FY_27_Ledger.xlsx"``).

        In the sandbox, XLSX files are pre-converted to CSV:
            ``FY_27_Ledger.xlsx`` → loadable as ``FY_27_Ledger.csv``

        So ``python_code`` must read the CSV form, even for an XLSX source:

            filename:    "FY_27_Ledger.xlsx"      (source name)
            python_code: pd.read_csv('FY_27_Ledger.csv', low_memory=False)
                                         ^^^ .csv, not .xlsx

        CSV files keep their name unchanged in the sandbox.

        Handling leading metadata rows
        ------------------------------
        Some exports have metadata rows above the real header. The tool
        response always includes a schema footer with a ready-to-use
        ``load:`` command that accounts for this — e.g.
        ``pd.read_csv('file.csv', skiprows=3, low_memory=False)``.
        **On any retry, use that exact load line verbatim** instead of
        guessing ``skiprows``.

        Best for: aggregations, filtering, trends, comparisons, statistics,
        charts. For simple factual lookups, use knowledge base search.

        Args:
            filename: Source filename from list_spreadsheets results. Use
                the original name (``.xlsx`` or ``.csv``), not the sandbox
                form.
            python_code: Python to execute. Load XLSX sources via
                ``pd.read_csv('<stem>.csv', ...)``. Available libraries:
                pandas, numpy, matplotlib, seaborn, openpyxl.
            output_filename: Optional PNG filename if generating a chart.
                Must end with ``.png``. Example: ``"chart.png"``.

        Returns:
            Analysis results as text (with a schema footer), and optionally
            a chart image.
        """
        from bedrock_agentcore.tools.code_interpreter_client import CodeInterpreter

        # 1. Validate Code Interpreter is available
        ci_id = _get_code_interpreter_id()
        if not ci_id:
            return {"content": [{"text": "❌ Code Interpreter is not configured. Contact your administrator."}], "status": "error"}

        # 2. Find the file in accessible sources
        file_info = _find_file(filename, assistant_id, session_id)
        if not file_info:
            return {"content": [{"text": f"❌ File '{filename}' not found or not accessible. Use list_spreadsheets to see available files."}], "status": "error"}

        # 3. Download from S3
        try:
            file_bytes = _download_file(file_info)
        except Exception as e:
            return {"content": [{"text": f"❌ Failed to download file: {e}"}], "status": "error"}

        # 4. Push file to Code Interpreter
        content_type = file_info.get("content_type", "")
        is_xlsx = "spreadsheetml" in content_type or filename.lower().endswith(".xlsx")

        region = os.getenv("AWS_REGION", "us-west-2")
        code_interpreter = CodeInterpreter(region)

        try:
            code_interpreter.start(identifier=ci_id)

            if is_xlsx:
                # Push XLSX as base64, decode in sandbox. Only the first sheet
                # is converted; if the workbook has multiple sheets we surface
                # a warning so the model can tell the user rather than silently
                # analyzing the wrong tab.
                import base64
                b64_content = base64.b64encode(file_bytes).decode("ascii")
                csv_filename = os.path.splitext(filename)[0] + ".csv"

                code_interpreter.invoke("writeFiles", {"content": [
                    {"path": "_encoded.b64", "text": b64_content},
                ]})
                bootstrap_code = f"""
import base64, io, csv
from openpyxl import load_workbook

with open('_encoded.b64', 'r') as f:
    raw = base64.b64decode(f.read())

wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
_active_sheet = wb.sheetnames[0]
ws = wb[_active_sheet]
with open('{csv_filename}', 'w', newline='') as out:
    writer = csv.writer(out)
    for row in ws.iter_rows(values_only=True):
        if all(cell is None for cell in row):
            continue
        writer.writerow([str(cell) if cell is not None else '' for cell in row])

# Emit the sheet inventory so the caller can warn about multi-sheet workbooks.
print({_SHEETS_MARKER!r})
print(f'active: {{_active_sheet}}')
print(f'all: {{wb.sheetnames}}')
print({_SHEETS_MARKER!r})
wb.close()
"""
                multi_sheet_note = ""
                resp = code_interpreter.invoke("executeCode", {"code": bootstrap_code, "language": "python", "clearContext": False})
                bootstrap_stdout = ""
                for event in resp.get("stream", []):
                    result = event.get("result", {})
                    if result.get("isError", False):
                        error_msg = _clean_stderr(result.get("structuredContent", {}).get("stderr", ""))
                        return {"content": [{"text": f"❌ Failed to convert XLSX in sandbox:\n```\n{error_msg}\n```"}], "status": "error"}
                    bootstrap_stdout += result.get("structuredContent", {}).get("stdout", "")

                # Parse the sheet inventory emitted by the bootstrap.
                if _SHEETS_MARKER in bootstrap_stdout:
                    try:
                        block = bootstrap_stdout.split(_SHEETS_MARKER)[1].strip()
                        active = ""
                        all_sheets: list[str] = []
                        for line in block.splitlines():
                            if line.startswith("active:"):
                                active = line.split(":", 1)[1].strip()
                            elif line.startswith("all:"):
                                # Parse the Python list literal ("['a', 'b']").
                                import ast
                                try:
                                    all_sheets = ast.literal_eval(line.split(":", 1)[1].strip())
                                except (ValueError, SyntaxError):
                                    all_sheets = []
                        if len(all_sheets) > 1:
                            others = [s for s in all_sheets if s != active]
                            multi_sheet_note = (
                                f"⚠ Workbook has {len(all_sheets)} sheets; analyzing only '{active}'. "
                                f"Other sheets not loaded: {', '.join(others[:5])}"
                                + (f" (+{len(others) - 5} more)" if len(others) > 5 else "")
                            )
                    except Exception as e:
                        logger.warning(f"Failed to parse XLSX sheet inventory: {e}")
            else:
                # CSV — push directly as text
                csv_filename = filename if filename.lower().endswith(".csv") else os.path.splitext(filename)[0] + ".csv"
                multi_sheet_note = ""
                try:
                    csv_text = file_bytes.decode("utf-8")
                except UnicodeDecodeError:
                    csv_text = file_bytes.decode("utf-8", errors="replace")
                code_interpreter.invoke("writeFiles", {"content": [{"path": csv_filename, "text": csv_text}]})

            # 5. Probe schema — separate exec so its output is isolated from user code.
            schema_preview = ""
            try:
                preview_resp = code_interpreter.invoke("executeCode", {
                    "code": _build_preview_code(csv_filename),
                    "language": "python",
                    "clearContext": False,
                })
                preview_stdout = ""
                for event in preview_resp.get("stream", []):
                    result = event.get("result", {})
                    if result.get("isError", False):
                        continue
                    preview_stdout += result.get("structuredContent", {}).get("stdout", "")
                schema_preview, _ = _extract_schema_preview(preview_stdout)
            except Exception as e:
                logger.warning(f"Schema preview failed for {csv_filename}: {e}")

            # 6. Execute user code
            response = code_interpreter.invoke("executeCode", {
                "code": python_code,
                "language": "python",
                "clearContext": False,
            })

            execution_output = ""
            for event in response.get("stream", []):
                result = event.get("result", {})
                if result.get("isError", False):
                    error_msg = _clean_stderr(result.get("structuredContent", {}).get("stderr", ""))
                    error_text = f"❌ Code execution failed:\n```\n{error_msg}\n```"

                    # Targeted hint for the most common wrong-filename error:
                    # the model wrote `pd.read_csv('FY_27_Ledger.xlsx', ...)`
                    # but in the sandbox the file lives as `FY_27_Ledger.csv`
                    # (see docstring: XLSX sources are pre-converted). Naming
                    # this out explicitly is much more effective than relying
                    # on the model to infer it from the schema footer.
                    if (
                        is_xlsx
                        and "FileNotFoundError" in error_msg
                        and filename in error_msg
                    ):
                        error_text += (
                            f"\n\n**Hint:** In the sandbox, the XLSX source "
                            f"`{filename}` is loaded as `{csv_filename}`. "
                            f"Retry with `pd.read_csv('{csv_filename}', "
                            f"low_memory=False)`."
                        )

                    if schema_preview:
                        # Drop the first_row dump on errors — the load line +
                        # column list is enough for the retry, first_row is
                        # ~1K tokens of bloat on a path that's already costing
                        # a round-trip.
                        trimmed_schema = _strip_first_row(schema_preview)
                        error_text += f"\n\nDataset info (use the `load:` line verbatim):\n```\n{trimmed_schema}\n```"
                    else:
                        error_text += f"\n\nTry: `pd.read_csv('{csv_filename}', low_memory=False)`"
                    if multi_sheet_note:
                        error_text += f"\n\n{multi_sheet_note}"
                    return {"content": [{"text": error_text}], "status": "error"}
                stdout = result.get("structuredContent", {}).get("stdout", "")
                if stdout:
                    execution_output += stdout

            # 7. Download chart if requested
            success_text = _truncate_output(execution_output) or "✅ Code executed successfully (no output)."
            if schema_preview:
                success_text = f"{success_text}\n\n---\nDataset: {schema_preview.splitlines()[0] if schema_preview else ''}"
            if multi_sheet_note:
                success_text = f"{success_text}\n{multi_sheet_note}"

            if output_filename and output_filename.endswith(".png"):
                try:
                    dl_response = code_interpreter.invoke("readFiles", {"paths": [output_filename]})
                    file_content = None
                    for event in dl_response.get("stream", []):
                        result = event.get("result", {})
                        if "content" in result:
                            for block in result["content"]:
                                if "data" in block:
                                    file_content = block["data"]
                                elif "resource" in block and "blob" in block["resource"]:
                                    file_content = block["resource"]["blob"]
                                if file_content:
                                    break
                        if file_content:
                            break

                    if file_content:
                        return {
                            "content": [
                                {"text": success_text},
                                {"image": {"format": "png", "source": {"bytes": file_content}}},
                            ],
                            "status": "success",
                        }
                except Exception as e:
                    logger.warning(f"Failed to download chart {output_filename}: {e}")

            return {
                "content": [{"text": success_text}],
                "status": "success",
            }

        finally:
            try:
                code_interpreter.stop()
            except Exception:
                pass

    return analyze_spreadsheet


def _find_file(filename: str, assistant_id: Optional[str], session_id: str) -> Optional[Dict[str, Any]]:
    """Find a file by name in accessible sources. Returns file info or None.

    Matches are tolerant to XLSX ↔ CSV aliasing: if the model asks for
    ``foo.csv`` but only ``foo.xlsx`` exists (because the sandbox converts
    XLSX → CSV and the model copied the sandbox name into the ``filename``
    param on retry), we treat them as the same file. Prevents the common
    round-trip loop where analyze_spreadsheet rejects a reasonable guess
    and forces the model to call list_spreadsheets (#206).
    """
    candidates: list[Dict[str, Any]] = []
    if assistant_id:
        candidates.extend(_get_kb_files(assistant_id))
    candidates.extend(_get_session_files(session_id))

    target_lower = filename.lower()
    target_stem, _ = os.path.splitext(target_lower)

    # First pass: exact match (case-insensitive).
    for f in candidates:
        if f["filename"].lower() == target_lower:
            return f

    # Second pass: same stem, tabular extension. Covers foo.csv -> foo.xlsx
    # and foo.xlsx -> foo.csv. Only applies to tabular files so we don't
    # accidentally alias foo.pdf to foo.docx.
    from apis.shared.files.models import is_tabular_file

    if target_stem and any(target_lower.endswith(ext) for ext in (".csv", ".xls", ".xlsx")):
        for f in candidates:
            cand_lower = f["filename"].lower()
            cand_stem, _ = os.path.splitext(cand_lower)
            if cand_stem == target_stem and is_tabular_file(f["filename"], f.get("content_type", "")):
                return f

    return None


def _download_file(file_info: Dict[str, Any]) -> bytes:
    """Download file bytes from S3."""
    region = os.environ.get("AWS_REGION", "us-west-2")
    s3 = boto3.client("s3", region_name=region)

    if file_info["source"] == "knowledge_base":
        bucket = os.environ.get("S3_ASSISTANTS_DOCUMENTS_BUCKET_NAME")
        if not bucket:
            raise ValueError("S3_ASSISTANTS_DOCUMENTS_BUCKET_NAME not configured")
    else:
        bucket = file_info.get("s3_bucket")
        if not bucket:
            raise ValueError("S3 bucket not found in file metadata")

    response = s3.get_object(Bucket=bucket, Key=file_info["s3_key"])
    return response["Body"].read()
